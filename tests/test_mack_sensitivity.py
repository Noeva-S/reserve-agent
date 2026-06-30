from __future__ import annotations

from io import BytesIO
import unittest

import pandas as pd
from openpyxl import load_workbook

from reserve_agent.agent.context_builder import build_chat_context
from reserve_agent.agent.chat_agent import answer_user_question
from reserve_agent.data.loader import DataQualityReport
from reserve_agent.exports import build_excel_download
from reserve_agent.models.reserving import run_reserving_models


class MackSensitivityTests(unittest.TestCase):
    def setUp(self) -> None:
        self.triangle = pd.DataFrame(
            {
                0: [1000.0, 1100.0, 1250.0, 1300.0],
                1: [1500.0, 1650.0, 1800.0, None],
                2: [1750.0, 1900.0, None, None],
                3: [1880.0, None, None, None],
            },
            index=[2020, 2021, 2022, 2023],
        )
        self.report = DataQualityReport(
            row_count=10,
            claim_count=4,
            accident_years=[2020, 2021, 2022, 2023],
            valuation_years=[2020, 2021, 2022, 2023],
            missing_values=0,
            negative_amount_cells=0,
            zero_claim_rows=0,
            notes=[],
        )

    def test_mack_outputs_and_sensitivities_are_available(self) -> None:
        outputs = run_reserving_models(self.triangle, expected_loss_ratio_value=0.72)

        self.assertIsNotNone(outputs.mack)
        self.assertIn("Mack Reserve", outputs.mack.columns)
        self.assertIn("Mack Standard Error", outputs.mack.columns)
        self.assertIn("Mack 95% Upper", outputs.mack.columns)
        self.assertGreaterEqual(outputs.diagnostics["total_mack_reserve"], 0.0)
        self.assertGreaterEqual(outputs.diagnostics["total_mack_standard_error"], 0.0)

        self.assertIsNotNone(outputs.expected_lr_sensitivity)
        self.assertIn("Expected Loss Ratio", outputs.expected_lr_sensitivity.columns)
        self.assertGreater(len(outputs.expected_lr_sensitivity), 1)

        self.assertIsNotNone(outputs.factor_sensitivity)
        self.assertIn("Factor Shock", outputs.factor_sensitivity.columns)
        self.assertGreater(len(outputs.factor_sensitivity), 1)

    def test_download_and_agent_context_include_mack_results(self) -> None:
        outputs = run_reserving_models(self.triangle, expected_loss_ratio_value=0.72)
        excel_bytes = build_excel_download(outputs, self.triangle, self.report)
        workbook = load_workbook(BytesIO(excel_bytes), read_only=True)

        self.assertIn("Mack Results", workbook.sheetnames)
        self.assertIn("ELR Sensitivity", workbook.sheetnames)
        self.assertIn("Factor Sensitivity", workbook.sheetnames)

        context = build_chat_context(self.report, outputs)
        self.assertTrue(context["capabilities"]["has_mack_result"])
        self.assertTrue(context["capabilities"]["has_uncertainty_metrics"])
        answer = answer_user_question("Mack 的 95% 区间是多少？", context)
        self.assertIn("95%", answer)
        self.assertIn("Mack", answer)


if __name__ == "__main__":
    unittest.main()
